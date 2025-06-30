#!/usr/bin/python3

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import datetime

read_wb = load_workbook(u'./견적양식.xlsx', data_only=True)
read_ws = read_wb["Sheet1"]

from PyInquirer import prompt

questions = [
    {
        'type': 'list',
        'name': 'my_company',
        'message': 'What\'s your company name',
        'choices': ['은파산업', '드론맵'],
	    'default': '은파산업'
    },
    {
        'type': 'input',
        'name': 'company',
        'message': 'What\'s company to give estimate '
    },
    {
        'type': 'input',
        'name': 'name',
        'message': 'What\'s name to give estimate '
    },
    {
        'type': 'list',
        'name': 'vat',
        'message': 'VAT included ? ',
        'choices': ['별도', '포함'],
	    'default': '별도'
    },
    {
        'type': 'input',
        'name': 'item1',
        'message': 'Item to give estimate '
    },
    {
        'type': 'input',
        'name': 'item1_count',
        'message': 'Count ? ',
        'filter': lambda val: int(val),
	    'default': '1'
    },
    {
        'type': 'input',
        'name': 'item1_price',
        'message': 'Price ? ',
        'filter': lambda val: int(val),
	    'default': '10000'
    },
    {
        'type': 'input',
        'name': 'item2',
        'message': 'More item to give estimate ',
    },
    {
        'type': 'input',
        'name': 'item2_count',
        'message': 'Count ? ',
        'filter': lambda val: int(val),
	    'default': '1',
    },
    {
        'type': 'input',
        'name': 'item2_price',
        'message': 'Price ? ',
        'filter': lambda val: int(val),
        'when': lambda answers: answers['item2'] != '',
	    'default': '10000'
    }
]

answers = prompt(questions)
#print(answers)

my_company = answers['my_company']
company = answers['company']
name = answers['name']

read_ws['B3'] = company
read_ws['D4'] = '{} {}'.format(company, name)

read_ws['D5'] = datetime.date.today()
read_ws['D5'].number_format = 'yyyy-mm-dd'
estimate_date = read_ws['D5'].value

read_ws['H4'] = my_company

item1 = answers['item1']
item1_count = answers['item1_count']
item1_price = answers['item1_price']

START_ROW_OF_ITEMS = 14

row = START_ROW_OF_ITEMS
read_ws['B{}'.format(row)] = 1
read_ws['C{}'.format(row)] = item1
read_ws['F{}'.format(row)] = item1_count
read_ws['G{}'.format(row)] = item1_price
read_ws['I{}'.format(row)] = '=F{}*G{}'.format(row, row)

if answers['item2'] != '':
  item2 = answers['item2']
  item2_count = answers['item2_count']
  item2_price = answers['item2_price']

  row = row + 1
  read_ws['B{}'.format(row)] = 2
  read_ws['C{}'.format(row)] = item2
  read_ws['F{}'.format(row)] = item2_count
  read_ws['G{}'.format(row)] = item2_price
  read_ws['I{}'.format(row)] = '=F{}*G{}'.format(row, row)

# 부가세
vat = answers['vat']
read_ws['B28'] = '공급 금액  (부가세 {})'.format(vat)

# Summary
read_ws['G28'] = '=SUM(I14:I27)'
read_ws['D10'] = '="일금"&NUMBERSTRING(G28,1)&"원정"&TEXT(G28,"(\#,###)")&" 부가세 {}"'.format(vat)

stamp = Image('{}.png'.format(my_company))
pixels = 2.5 * 96.0 / 2.54 # 2.5cm to pixels
stamp.height = pixels 
stamp.width = pixels 
stamp.anchor = 'J6'
read_ws.add_image(stamp)

saved_file = './{}_{}_견적서.xlsx'.format(company, estimate_date)
read_wb.save(saved_file)
read_wb.close()

# Print saved results

# for linux
result_wb = load_workbook(saved_file, data_only=True)
result_ws = result_wb["Sheet1"]
estimate_list = []

row = 10 
for i in range(14, 14+row):
    items = []
    for j in range(1, 15):
        items.append(result_ws.cell(i, j).value)
    _items = list(filter(None, items))
    if _items:
        estimate_list.append(_items)

print(estimate_list)

# for mac
#import xlwings as xw
#result_wb = xw.Book(saved_file)
#result_ws = result_wb.sheets["Sheet1"]
#
#estimate_list = result_ws.range('B14:I20').value
#print(result_ws.range('D10').value)
