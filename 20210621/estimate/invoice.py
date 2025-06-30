#!/usr/bin/python3

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import datetime
import configparser
from PyInquirer import prompt

questions = [
    {
        'type': 'list',
        'name': 'my_company',
        'message': 'What\'s your company name',
        'choices': ['은파산업', '드론맵'],
        'default': '은파산업'
    },
    {
        'type': 'input',
        'name': 'company',
        'message': 'What\'s company to give invoice '
    },
    {
        'type': 'list',
        'name': 'vat',
        'message': 'VAT included ? ',
        'choices': ['별도', '포함'],
        'default': '별도'
    },
    {
        'type': 'input',
        'name': 'item1',
        'message': 'Item to give invoice '
    },
    {
        'type': 'input',
        'name': 'item1_count',
        'message': 'Count ? ',
        'filter': lambda val: int(val),
        'default': '1'
    },
    {
        'type': 'input',
        'name': 'item1_price',
        'message': 'Price ? ',
        'filter': lambda val: int(val),
	    'default': '10000'
    },
    {
        'type': 'input',
        'name': 'item2',
        'message': 'More item to give invoice ',
    },
    {
        'type': 'input',
        'name': 'item2_count',
        'message': 'Count ? ',
        'filter': lambda val: int(val),
        'when': lambda answers: answers['item2'] != '',
        'default': '1',
    },
    {
        'type': 'input',
        'name': 'item2_price',
        'message': 'Price ? ',
        'filter': lambda val: int(val),
        'when': lambda answers: answers['item2'] != '',
        'default': '10000'
    }
]

print('Generate invoice')

answers = prompt(questions)
#print(answers)

_my_company = answers['my_company']
_company = answers['company']

read_wb = load_workbook(u'./은파산업_거래명세양식.xlsx', data_only=True)
read_ws = read_wb["Sheet1"]

config = configparser.ConfigParser()
config.read('invoice.ini')
my_company = config[_my_company]
company = config[_company]

read_ws['A2'] = datetime.date.today()
read_ws['A2'].number_format = 'yyyy-mm-dd'
invoice_date = str(read_ws['A2'].value)

read_ws['E3'] = my_company['business_number']
read_ws['E5'] = my_company['name']
read_ws['M5'] = my_company['ceo']
read_ws['E7'] = my_company['address']
read_ws['E9'] = my_company['type']
read_ws['L9'] = my_company['business_item']

stamp = Image('{}.png'.format(_my_company))
pixels = 1.8 * 96.0 / 2.54 # 2.5cm to pixels
stamp.height = pixels 
stamp.width = pixels 
stamp.anchor = 'O4'
read_ws.add_image(stamp)

read_ws['U3'] = company['business_number']
read_ws['U5'] = company['name']
read_ws['AC5'] = company['ceo']
read_ws['U7'] = company['address']
read_ws['U9'] = company['type']
read_ws['AB9'] = company['business_item']

item_month = invoice_date.split('-')[1]
item_day = invoice_date.split('-')[2]
item1 = answers['item1']
item1_count = answers['item1_count']
item1_price = answers['item1_price']

START_ROW_OF_ITEMS = 12

row = START_ROW_OF_ITEMS
read_ws['A{}'.format(row)] = item_month
read_ws['B{}'.format(row)] = item_day
read_ws['C{}'.format(row)] = item1
read_ws['M{}'.format(row)] = item1_count
read_ws['O{}'.format(row)] = item1_price
read_ws['T{}'.format(row)] = '=M{}*O{}'.format(row, row)

if answers['item2'] != '':
  item2 = answers['item2']
  item2_count = answers['item2_count']
  item2_price = answers['item2_price']

  row = str(int(row) + 1)

  read_ws['A{}'.format(row)] = item_month
  read_ws['B{}'.format(row)] = item_day
  read_ws['C{}'.format(row)] = item2
  read_ws['M{}'.format(row)] = item2_count
  read_ws['O{}'.format(row)] = item2_price
  read_ws['T{}'.format(row)] = '=M{}*O{}'.format(row, row)

# 부가세
vat = answers['vat']
#read_ws['B28'] = '부가세 {}'.format(vat)
read_ws['C16'] = '=SUM(T12:Y15)'
read_ws['J16'] = '=SUM(Z12:AD15)'
read_ws['P16'] = '=C16+J16'

now = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
saved_file = './results/{}_{}_{}_거래명세표.xlsx'.format(_company, now, _my_company)
read_wb.save(saved_file)
read_wb.close()

# Print saved results
result_wb = load_workbook(saved_file, data_only=True)
result_ws = result_wb["Sheet1"]
invoice_list = []

row = 4
for i in range(12, 12+row):
    items = []
    for j in range(1, 40):
        items.append(result_ws.cell(i, j).value)
    _items = list(filter(None, items))
    if _items:
        invoice_list.append(_items)

print(invoice_list)
