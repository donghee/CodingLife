#!/usr/bin/python3

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import datetime
import configparser
from PyInquirer import prompt

questions = [
    {
        'type': 'list',
        'name': 'my_company',
        'message': 'What\'s your company name',
        'choices': ['은파산업', '드론맵', '미래전자기술믹서', '이재경스튜디오'],
	    'default': '은파산업'
    },
    {
        'type': 'input',
        'name': 'company',
        'message': 'What\'s company to give estimate '
    },
    {
        'type': 'input',
        'name': 'name',
        'message': 'What\'s name to give estimate '
    },
    {
        'type': 'list',
        'name': 'vat',
        'message': 'VAT included ? ',
        'choices': ['별도', '포함'],
	    'default': '별도'
    },
    {
        'type': 'input',
        'name': 'item1',
        'message': 'Item to give estimate '
    },
    {
        'type': 'input',
        'name': 'item1_count',
        'message': 'Count ? ',
        'filter': lambda val: int(val),
	    'default': '1'
    },
    {
        'type': 'input',
        'name': 'item1_price',
        'message': 'Price ? ',
        'filter': lambda val: int(val),
	    'default': '10000'
    },
    {
        'type': 'input',
        'name': 'item2',
        'message': 'More item to give estimate ',
    },
    {
        'type': 'input',
        'name': 'item2_count',
        'message': 'Count ? ',
        'filter': lambda val: int(val),
        'when': lambda answers: answers['item2'] != '',
	    'default': '1',
    },
    {
        'type': 'input',
        'name': 'item2_price',
        'message': 'Price ? ',
        'filter': lambda val: int(val),
        'when': lambda answers: answers['item2'] != '',
	    'default': '10000'
    }
]

print('Generate estimation')

answers = prompt(questions)
#print(answers)

my_company = answers['my_company']
company = answers['company']
name = answers['name']

config = configparser.ConfigParser()
config.read('estimate.ini')
address = config[my_company]

read_wb = load_workbook(u'./{}'.format(config[my_company]['template']), data_only=True)
read_ws = read_wb["Sheet1"]

read_ws[address['company']] = company
read_ws[address['company_name']] = '{} {}'.format(company, name)

read_ws[address['today']] = datetime.date.today()
read_ws[address['today']].number_format = 'yyyy-mm-dd'
estimate_date = read_ws[address['today']].value

read_ws[address['my_company']] = my_company

item1 = answers['item1']
item1_count = answers['item1_count']
item1_price = answers['item1_price']

START_ROW_OF_ITEMS = address['item_row_start']

row = START_ROW_OF_ITEMS
read_ws['{}{}'.format(address['item_number'], row)] = 1
read_ws['{}{}'.format(address['item'] ,row)] = item1
read_ws['{}{}'.format(address['item_count'], row)] = item1_count
read_ws['{}{}'.format(address['item_price'], row)] = item1_price
read_ws['{}{}'.format(address['item_sum'], row)] = '={}{}*{}{}'.format(address['item_count'], row, address['item_price'], row)

if answers['item2'] != '':
  item2 = answers['item2']
  item2_count = answers['item2_count']
  item2_price = answers['item2_price']

  row = int(row) + 1
  read_ws['{}{}'.format(address['item_number'], row)] = 2
  read_ws['{}{}'.format(address['item'] ,row)] = item2
  read_ws['{}{}'.format(address['item_count'], row)] = item2_count
  read_ws['{}{}'.format(address['item_price'], row)] = item2_price
  read_ws['{}{}'.format(address['item_sum'], row)] = '={}{}*{}{}'.format(address['item_count'], row, address['item_price'], row)

# 부가세
vat = answers['vat']
read_ws[address['price']] = '공급 금액  (부가세 {})'.format(vat)

# Summary
read_ws[address['total_price']] = '=SUM({})'.format(address['items_sum'])
read_ws[address['hangul_total_price']] = '="일금"&NUMBERSTRING({},1)&"원정"&TEXT({},"(\#,###)")&" 부가세 {}"'.format(address['total_price'], address['total_price'], vat)

stamp = Image('{}.png'.format(my_company))
stamp.anchor = address['stamp_anchor']
pixels = float(address['stamp_size']) * 96.0 / 2.54 # (stamp_size) cm to pixels
stamp.height = pixels 
stamp.width = pixels 
read_ws.add_image(stamp)

saved_file = './results/{}_{}_{}_견적서.xlsx'.format(company, estimate_date, my_company)
read_wb.save(saved_file)
read_wb.close()

# Print saved results

# for linux
result_wb = load_workbook(saved_file, data_only=True)
result_ws = result_wb["Sheet1"]
estimate_list = []

row = 10 
for i in range(14, 14+row):
    items = []
    for j in range(1, 15):
        items.append(result_ws.cell(i, j).value)
    _items = list(filter(None, items))
    if _items:
        estimate_list.append(_items)

print(estimate_list)

# for mac
#import xlwings as xw
#result_wb = xw.Book(saved_file)
#result_ws = result_wb.sheets["Sheet1"]
#
#estimate_list = result_ws.range('B14:I20').value
#print(result_ws.range('D10').value)
